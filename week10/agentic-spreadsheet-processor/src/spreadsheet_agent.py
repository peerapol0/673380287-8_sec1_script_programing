import logging
import json
import shutil
from datetime import datetime

import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.cell import range_boundaries


logger = logging.getLogger(__name__)


class SpreadsheetAgent:
    def __init__(self, config: dict):
        self.config = config

    def run(self) -> None:
        started_at = datetime.now()
        audit = {
            "started_at": started_at.isoformat(timespec="seconds"),
            "status": "running",
            "input_file": self.config["input_file"],
            "output_file": self.config["output_file"],
            "tasks": [],
        }
        output_file = self.config["output_file"]
        try:
            shutil.copyfile(self.config["input_file"], output_file)
            workbook = openpyxl.load_workbook(output_file)
            for task in self.config.get("tasks", []):
                task_type = task.get("type")
                if task_type == "copy_data":
                    self._copy_data(workbook, task)
                elif task_type == "calculate_column":
                    self._calculate_column(workbook, task)
                elif task_type == "conditional_format":
                    self._conditional_format(workbook, task)
                elif task_type == "chart":
                    self._add_chart(workbook, task)
                else:
                    raise ValueError(f"Unsupported task type: {task_type}")
                audit["tasks"].append({"type": task_type, "status": "success"})
            audit["finished_at"] = datetime.now().isoformat(timespec="seconds")
            workbook.save(output_file)
            audit["status"] = "success"
            audit["sheet_names"] = workbook.sheetnames
            audit["chart_count"] = sum(len(sheet._charts) for sheet in workbook.worksheets)
            self._add_audit_sheet(workbook, audit)
            workbook.save(output_file)
            logger.info("Spreadsheet saved to %s", output_file)
        except Exception as error:
            audit["status"] = "failed"
            audit["error"] = str(error)
            raise
        finally:
            audit["finished_at"] = datetime.now().isoformat(timespec="seconds")
            self._write_audit_log(audit)

    def _write_audit_log(self, audit: dict) -> None:
        audit_path = self.config.get("audit_log")
        if not audit_path:
            return
        with open(audit_path, "w", encoding="utf-8") as audit_file:
            json.dump(audit, audit_file, indent=2, ensure_ascii=False)
        logger.info("Audit log saved to %s", audit_path)

    @staticmethod
    def _add_audit_sheet(workbook, audit: dict) -> None:
        sheet_name = "Audit Log"
        if sheet_name in workbook.sheetnames:
            del workbook[sheet_name]
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(["Field", "Value"])
        sheet.append(["Started At", audit["started_at"]])
        sheet.append(["Finished At", audit["finished_at"]])
        sheet.append(["Status", audit["status"]])
        sheet.append(["Input File", audit["input_file"]])
        sheet.append(["Output File", audit["output_file"]])
        sheet.append([])
        sheet.append(["Task #", "Task Type", "Status"])
        for task_number, task in enumerate(audit["tasks"], start=1):
            sheet.append([task_number, task["type"], task["status"]])
        sheet.column_dimensions["A"].width = 18
        sheet.column_dimensions["B"].width = 85
        sheet.column_dimensions["C"].width = 15

    @staticmethod
    def _copy_data(workbook, task):
        source = workbook[task["source_sheet"]]
        if task["dest_sheet"] in workbook.sheetnames:
            del workbook[task["dest_sheet"]]
        destination = workbook.create_sheet(task["dest_sheet"])
        for row in source.iter_rows():
            for cell in row:
                destination[cell.coordinate] = cell.value

    @staticmethod
    def _calculate_column(workbook, task):
        sheet = workbook[task["sheet"]]
        column = task["target_column"]
        sheet[f"{column}1"] = task["header"]
        for row in range(task["start_row"], sheet.max_row + 1):
            cell = sheet[f"{column}{row}"]
            cell.value = task["formula"].format(row=row)
            if task.get("number_format"):
                cell.number_format = task["number_format"]

    @staticmethod
    def _conditional_format(workbook, task):
        sheet = workbook[task["sheet"]]
        if task["rule_type"] == "color_scale":
            rule = ColorScaleRule(start_type="min", start_color=task["min_color"], mid_type="percentile", mid_value=50, mid_color=task["mid_color"], end_type="max", end_color=task["max_color"])
        elif task["rule_type"] == "expression":
            rule = FormulaRule(formula=[task["formula"]], fill=PatternFill("solid", fgColor=task["fill_color"]), font=Font(color=task.get("font_color", "000000")))
        else:
            raise ValueError(f"Unsupported conditional format: {task['rule_type']}")
        sheet.conditional_formatting.add(task["range"], rule)

    @staticmethod
    def _add_chart(workbook, task):
        sheet = workbook[task["sheet"]]
        chart = BarChart()
        chart.title = task.get("title", "")
        chart.x_axis.title = task.get("x_axis_title", "")
        chart.y_axis.title = task.get("y_axis_title", "")
        data_reference = SpreadsheetAgent._reference(sheet, task["data_range"])
        chart.add_data(data_reference, titles_from_data=True)
        category_reference = SpreadsheetAgent._reference(sheet, task["category_range"])
        chart.set_categories(category_reference)
        sheet.add_chart(chart, task["top_left_cell"])

    @staticmethod
    def _reference(sheet, cell_range):
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
        return Reference(sheet, min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row)