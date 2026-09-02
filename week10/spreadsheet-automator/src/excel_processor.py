# spreadsheet-automator/src/excel_processor.py
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


class ExcelProcessor:
    """
    A class to encapsulate common Excel processing functionalities using openpyxl.
    """

    def __init__(self):
        pass

    def create_workbook(self, filename="new_workbook.xlsx"):
        """Creates a new Excel workbook."""
        wb = openpyxl.Workbook()
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        print(f"New workbook '{filename}' created.")
        return wb

    def load_workbook(self, filename):
        """Loads an existing Excel workbook."""
        try:
            wb = openpyxl.load_workbook(filename)
            print(f"Workbook '{filename}' loaded successfully.")
            return wb
        except FileNotFoundError:
            print(f"Error: Workbook '{filename}' not found.")
            return None
        except Exception as e:
            print(f"Error loading workbook '{filename}': {e}")
            return None

    def save_workbook(self, workbook, filename):
        """Saves a workbook to a specified filename."""
        try:
            workbook.save(filename)
            print(f"Workbook saved to '{filename}'.")
            return True
        except Exception as e:
            print(f"Error saving workbook to '{filename}': {e}")
            return False

    def process_sales_data(self, input_filename, output_filename):
        """
        Loads sales data, calculates totals, and writes to a new formatted Excel file.
        Assumes input columns: Product Name, Quantity, Unit Price.
        """
        input_wb = self.load_workbook(input_filename)
        if not input_wb:
            return False

        input_sheet = input_wb.active  # Assumes data is in the active sheet

        # Create a new workbook for output
        output_wb = self.create_workbook()
        output_sheet = output_wb.create_sheet("Sales Report", 0)

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")  # Green
        total_row_font = Font(bold=True, color="000000")
        total_row_fill = PatternFill(start_color="FFEB3B", end_color="FFEB3B", fill_type="solid")  # Yellow
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )
        currency_format = '#,##0.00'  # Standard currency format

        # Write headers to output sheet
        headers = ["Product Name", "Quantity", "Unit Price", "Total Price"]
        for col_num, header in enumerate(headers, 1):
            cell = output_sheet.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

        grand_total = 0
        output_row = 2  # Start writing data from row 2

        # Iterate through rows in input sheet, skipping header
        for row_index, row in enumerate(input_sheet.iter_rows(min_row=2, values_only=True), 2):
            if not any(row):  # Skip empty rows
                continue

            product_name = row[0] if len(row) > 0 else None
            quantity = row[1] if len(row) > 1 else 0
            unit_price = row[2] if len(row) > 2 else 0.0

            # Basic type conversion and error handling
            try:
                quantity = int(quantity)
                unit_price = float(unit_price)
            except (ValueError, TypeError):
                print(f"Warning: Skipping row {row_index} due to invalid number format: {row}")
                continue

            total_price = quantity * unit_price
            grand_total += total_price

            # Write data to output sheet
            output_sheet.cell(row=output_row, column=1, value=product_name).border = thin_border
            output_sheet.cell(row=output_row, column=2, value=quantity).border = thin_border
            output_sheet.cell(row=output_row, column=3, value=unit_price).border = thin_border
            total_price_cell = output_sheet.cell(row=output_row, column=4, value=total_price)
            total_price_cell.number_format = currency_format
            total_price_cell.border = thin_border

            output_row += 1

        # Write grand total
        output_sheet.merge_cells(start_row=output_row, start_column=1, end_row=output_row, end_column=3)
        grand_total_label_cell = output_sheet.cell(row=output_row, column=1, value="Grand Total:")
        grand_total_label_cell.font = total_row_font
        grand_total_label_cell.fill = total_row_fill
        grand_total_label_cell.alignment = Alignment(horizontal='right', vertical='center')  # Align label right

        grand_total_value_cell = output_sheet.cell(row=output_row, column=4, value=grand_total)
        grand_total_value_cell.font = total_row_font
        grand_total_value_cell.fill = total_row_fill
        grand_total_value_cell.number_format = currency_format
        grand_total_value_cell.border = thin_border

        # Adjust column widths for better readability
        for col_num in range(1, len(headers) + 1):
            max_length = 0
            column_letter = get_column_letter(col_num)
            for cell in output_sheet[column_letter]:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            adjusted_width = (max_length + 2)  # Add a little padding
            if adjusted_width > 0:  # Avoid setting zero width
                output_sheet.column_dimensions[column_letter].width = adjusted_width

        return self.save_workbook(output_wb, output_filename)
